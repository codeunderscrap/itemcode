"""Builds Sprint_Plan.xlsx (v2) for Item Code Studio — single-sitting delivery plan."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

F = "Arial"
NAVY = "001B2E"; STEEL = "3B6E93"; TEAL = "04AED1"
LGREY = "F2F2F2"; BAND = "F7FAFC"; GREEN = "E2EFDA"; DONEG = "C6E0B4"
AMBER = "FFF2CC"; RED = "FCE4EC"; ICE = "E8F4F8"

H = Font(name=F, sz=10, bold=True, color="FFFFFF")
B = Font(name=F, sz=10, bold=True)
N = Font(name=F, sz=10)
S = Font(name=F, sz=9, color="595959")
TITLE = Font(name=F, sz=14, bold=True, color=NAVY)
SUB = Font(name=F, sz=11, bold=True, color=STEEL)
thin = Side(style="thin", color="D9D9D9")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center")


AGENT = {
 "Z1":"0 Foundation","Z2":"0 Foundation","Z3":"0 Foundation","Z4":"0 Foundation",
 "K1":"0 Foundation","K2":"A Public","A1":"A Public","A2":"A Public","A3":"A Public","A4":"A Public",
 "A5":"B Auth","A6":"B Auth","A7":"B Auth","A8":"B Auth","A9":"B Auth","A10":"B Auth","A11":"B Auth","S1":"B Auth",
 "D1":"C Matching","D2":"C Matching","D3":"C Matching","D4":"C Matching","D5":"C Matching",
 "D6":"C Matching","D7":"C Matching","D8":"C Matching","D9":"C Matching","D10":"C Matching","D11":"C Matching",
 "C1":"D Engine","C2":"D Engine","C3":"D Engine","C4":"D Engine","C5":"D Engine","C6":"D Engine",
 "C7":"D Engine","C8":"D Engine","C9":"D Engine","K4":"D Engine","K11":"D Engine",
 "C10":"E Create","F8":"E Create","K3":"E Create","K5":"E Create","B5":"E Create",
 "F1":"E Create","F2":"E Create","F6":"E Create","F7":"E Create","F3":"E Create","F4":"E Create","F5":"E Create",
 "K6":"F Master","K7":"F Master","K8":"F Master","K9":"F Master","K10":"F Master","K12":"F Master",
 "G1":"F Master","G2":"F Master","G3":"F Master","G4":"F Master","G5":"F Master","G6":"F Master",
 "H1":"G ERPNext","H2":"G ERPNext","H3":"G ERPNext","H4":"G ERPNext","H5":"G ERPNext","H6":"G ERPNext",
 "H7":"G ERPNext","H8":"G ERPNext","H9":"G ERPNext","H10":"G ERPNext","H11":"G ERPNext",
 "H12":"G ERPNext","H13":"G ERPNext","H14":"G ERPNext",
 "B1":"H Deploy","B2":"H Deploy","B3":"H Deploy","B4":"H Deploy","B6":"H Deploy","B7":"H Deploy",
 "J1":"H Deploy","J2":"H Deploy","J4":"H Deploy","J5":"H Deploy","J6":"H Deploy","J7":"H Deploy",
 "J8":"H Deploy","J9":"H Deploy","J10":"H Deploy",
 "I1":"— Anuraag","I2":"— Anuraag","I3":"— Anuraag","I4":"— Anuraag","I5":"— Anuraag","I6":"— Anuraag",
}

wb = openpyxl.Workbook()

# ID, Epic, Story, Acceptance criteria, Priority, Points, Run, Status, Depends, Notes
ST = [
 # ---------------------------------------------------------------- RUN 0  FOUNDATION
 ("Z1","E0 Brand & Shell","Schema migration","app_user, session, item_version, item_vacancy, llm_cache and sync_log created idempotently. Seeded data survives; counts unchanged after rebuild.","Must",8,"R1","Not started","","Agent 0 only - nobody else touches db.py"),
 ("Z2","E2 API v1","Split server.py into route modules","routes/ with public, auth, create, master, erp, meta. server.py becomes a thin dispatcher under 120 lines with the two response envelopes.","Must",8,"R1","Not started","","Stops 8 agents colliding in one file"),
 ("Z3","E0 Brand & Shell","Settings store and ledger config","db.get_setting / set_setting; config gains ledger.mode host|client. Secrets live in settings, never config.json.","Must",3,"R1","Not started","Z1",""),
 ("Z4","E0 Brand & Shell","Theme file and logo asset","web/theme.css with the MiniMines palette as variables, light variant, Barlow. minimines-logo.svg copied from ATT_Platform.","Must",2,"R1","Not started","","Agent A builds on these variables"),
 # ---------------------------------------------------------------- RUN 1
 ("K1","E0 Brand & Shell","Apply the MiniMines theme","Palette from m-mines.com (#001b2e navy, #04aed1 teal, #3b6e93 steel), Barlow typeface, minimines-logo.svg in the header. Light and dark both legible.","Must",5,"R1","Not started","","Reused from ATT_Platform, not reinvented"),
 ("K2","E0 Brand & Shell","Remove the stats block","The counts panel in the bottom-left corner is gone. Nothing shifts or leaves a gap.","Must",1,"R1","Not started","K1","Your instruction"),
 ("A1","E1 Access & Identity","Split landing page","One page, two halves: Decoder left, Directory right, Log in top-right. No credentials needed to use either half.","Must",8,"R1","Not started","K1","The main page for everyone"),
 ("A2","E1 Access & Identity","Decoder half","Paste any code, get head, sub-head, group, every spec with its per-group label, vendor. 00 reads 'not applicable'. Malformed input explains why.","Must",3,"R1","Not started","A1",""),
 ("A4","E1 Access & Identity","Directory half","Search 1,947 items by code, name or description. Row opens a read-only detail card.","Must",5,"R1","Not started","A1",""),
 ("A3","E1 Access & Identity","Dictionary browse","Search 889 groups; open one to see its spec slots and values. Reachable from the directory half.","Must",3,"R1","Not started","A4",""),
 ("A8","E1 Access & Identity","Server-side write protection","Every mutating endpoint returns 401 to an unauthenticated caller invoked directly, not merely hidden in the interface.","Must",3,"R1","Not started","","The actual boundary"),
 # ---------------------------------------------------------------- RUN 2
 ("A5","E1 Access & Identity","Login","Username and password against the user table, scrypt hashed. Bad credentials give one generic message.","Must",5,"R2","Not started","A8","Simple by instruction"),
 ("A6","E1 Access & Identity","Session handling","HMAC-signed HttpOnly cookie, 12h expiry, logout clears it. Expiry returns to the landing page, never a blank screen.","Must",3,"R2","Not started","A5",""),
 ("A9","E1 Access & Identity","Attribution from session","created_by comes from the session. The forgeable X-User header is deleted from the codebase.","Must",2,"R2","Not started","A6","Closes a real hole"),
 ("A7","E1 Access & Identity","Creator shell","After login: Create, Item master, Activity. The landing page stays reachable while logged in.","Must",3,"R2","Not started","A6",""),
 ("A10","E1 Access & Identity","User management","manage.py adduser / disable / resetpw. At least one active creator must always remain.","Must",3,"R2","Not started","A5","CLI is enough at this size"),
 ("S1","E1 Access & Identity","Settings screen with the API key","Admin-only. match.mode fuzzy|llm toggle, provider, API key as a write-only password field, model, threshold, erp flags, sync times. Key never returned to the client. With no key the app runs fuzzy-only and says so.","Must",8,"R2","Not started","A5","Anuraag pastes the key himself"),
 ("A11","E1 Access & Identity","Change own password","A logged-in creator can change their password; the old one is required.","Should",2,"R2","Not started","A5",""),
 # ---------------------------------------------------------------- RUN 3
 ("C10","E3 Code Engine","Create screen with the three phases","Paste or upload, then every line shows what phase 1 found, which group phase 2 chose, and what phase 3 read into each slot.","Must",5,"R3","Not started","A7","Engine exists; this is the gated screen"),
 ("F8","E5 Invoice & OCR","Invoice upload wired into Create","Drop a PDF, scan, photo, Excel or CSV and get one row per line item, each resolving independently.","Must",5,"R3","Not started","C10","Engine exists"),
 ("K3","E6 Master & Restructure","Conditional cascading dropdowns","Choosing a head filters sub-heads to that head. Choosing a sub-head filters groups to that sub-head only. Choosing a group re-renders its spec slots with that group's labels.","Must",8,"R3","Not started","C10","Core of your edit requirement"),
 ("K4","E6 Master & Restructure","Live code recomputation","Any change to head, sub-head, group, spec or vendor recomputes the code immediately, obeying the interior-00 and trailing-truncate rules.","Must",5,"R3","Not started","K3",""),
 ("K11","E3 Code Engine","Item-level number reservation","Moving an item to another group frees its old position under that head and sub-head into the vacancy ledger. New arrivals take a fresh number; the reserved one is released only to a later item with the same spec tuple, and dropped when the sub-head is revoked.","Must",8,"R3","Not started","K4","The logic you said was missing"),
 ("K12","E6 Master & Restructure","Vacancy visibility","Reserved numbers at both group and item level are listed with what freed them and what would claim them.","Should",3,"R4","Not started","K11",""),
 ("K5","E6 Master & Restructure","Edit before submit","Every proposed code carries an Edit control. Nothing is written until Submit.","Must",3,"R3","Not started","K4",""),
 ("B5","E2 API v1","Idempotent commit","A commit carries a client token; replaying it returns the original code instead of issuing a second.","Must",3,"R3","Not started","K5","Guards double-click"),
 ("C8","E3 Code Engine","Concurrent issuance safety","Two creators submitting in the same second cannot receive the same code. Proven by a concurrent test, not by argument.","Must",5,"R5","Not started","B5",""),
 # ---------------------------------------------------------------- RUN 4
 ("K6","E6 Master & Restructure","Item master as an editable directory","Same rows and columns as the public directory, every field editable in place, with the same cascading dropdowns.","Must",8,"R4","Not started","K3","Your instruction"),
 ("K7","E6 Master & Restructure","Revision history","Every save snapshots the whole row with who, when and what changed. Nothing is overwritten in place.","Must",8,"R4","Not started","K6","The version-control requirement"),
 ("K8","E6 Master & Restructure","Revert to a previous version","Any earlier version can be restored in one click. The revert is itself a new version, so it can be undone too.","Must",5,"R4","Not started","K7",""),
 ("K9","E6 Master & Restructure","Activity feed with diffs","Activity shows field-level before and after, filterable by person, item and date, with Revert on each entry.","Must",5,"R4","Not started","K7",""),
 ("G3","E6 Master & Restructure","Move, merge, retire, rename","Each shows a full impact preview listing what re-codes and what stays frozen.","Must",8,"S0","Done","","Built and tested on group Tissue"),
 ("K10","E6 Master & Restructure","Revert respects frozen codes","A revert never rewrites a code that is live in ERPNext. It restores fields and says plainly what it could not restore.","Must",5,"R4","Not started","K8","Where version control meets freeze-on-first-use"),
 ("G5","E6 Master & Restructure","Excel export","Item_Master, Item_Code_Master, Code_Mapping and Audit_Trail in one workbook.","Must",5,"S0","Done","",""),
 # ---------------------------------------------------------------- RUN 5
 ("B1","E2 API v1","Version and freeze the contract","Everything served under /api/v1 exactly as the PDR documents it.","Must",5,"R5","Not started","A8","So the larger system can build against it"),
 ("B2","E2 API v1","Uniform error envelope","Every failure returns ok:false with a stable machine code. No stack traces reach the client.","Must",3,"R5","Not started","B1",""),
 ("B3","E2 API v1","Self-served API docs","GET /api/docs lists every endpoint with request and response examples.","Should",3,"R5","Not started","B1",""),
 ("B4","E2 API v1","Pagination and filtering","List endpoints take q, limit, offset, status and return total alongside rows.","Should",2,"R5","Not started","B1",""),
 ("B6","E2 API v1","Public endpoint hardening","Payload caps and a simple per-IP rate limit on public reads.","Should",2,"R5","Not started","A8",""),
 ("B7","E2 API v1","Smoke suite","A stdlib script exercises every endpoint including auth failures and runs in under a minute.","Must",5,"R5","Not started","B2",""),
 ("J7","E9 Deploy & Handover","Deploy to the internal server","Runs on the nominated machine, reachable by hostname on the LAN, starts on boot, survives a restart.","Must",5,"R5","Not started","B7","Internal grade, not production grade"),
 ("J9","E9 Deploy & Handover","Client ledger mode","A creator install in client mode reads and writes the ledger on the LAN host. Local OCR, matching and LLM stay local. Unreachable host gives a plain message, not a stack trace.","Must",8,"R5","Not started","Z3","Keeps codes unique across installs"),
 ("J10","E9 Deploy & Handover","Local installer","Copies the app, writes config, creates shortcuts, checks for Python, optional start-on-boot for the host. No admin rights, no pip, no internet.","Must",8,"R5","Not started","J9","The app is installed locally now"),
 ("J2","E9 Deploy & Handover","Automated backup","Daily copy of itemcode.db plus a weekly Excel export, both written off the host machine.","Must",3,"R5","Not started","J7",""),
 ("J8","E9 Deploy & Handover","Seed and go-live check","Fresh seed from the source workbooks on the server; counts match; a code is issued end to end.","Must",3,"R5","Not started","J7",""),
 # ---------------------------------------------------------------- ALREADY BUILT
 ("C1","E3 Code Engine","Code grammar assembly","Interior gap 00, trailing gaps dropped, vendor forces 17. Only 7/9/11/13/15/17 accepted.","Must",8,"S0","Done","","Verified against RMBS0010206100007"),
 ("C2","E3 Code Engine","Three-phase resolver","Phase 1 stops on an existing code; phase 2 resolves head/sub-head/group; phase 3 resolves specs and vendor.","Must",13,"S0","Done","C1",""),
 ("C3","E3 Code Engine","Prefix collision ladder","A taken 4-letter prefix walks to the next free combination, checked against groups and live ERP codes.","Must",3,"S0","Done","C1","CEEU preserved as seeded"),
 ("C4","E3 Code Engine","Per-sub-head numbering and vacancies","Numbers restart per sub-head. A moved group's number is parked, released only to a >=88% name match.","Must",5,"S0","Done","C1","Your rule"),
 ("C5","E3 Code Engine","Freeze on first use","A code live in ERPNext is never rewritten by a structural change; it is kept and flagged stale.","Must",5,"S0","Done","C4","Needs your confirmation"),
 ("C6","E3 Code Engine","Decoder with per-group labels","Decode reports each slot's meaning for that group, and 00 as not applicable.","Must",3,"S0","Done","C1",""),
 ("C7","E3 Code Engine","Ambiguous spec is a question","A declared slot with no determinable value blocks submission; it is never auto-numbered.","Must",3,"S0","Done","C2","Caught a real collision"),
 ("D1","E4 Matching & LLM","Deterministic normaliser","Units, casing, punctuation, synonyms and noise words folded before scoring.","Must",5,"S0","Done","",""),
 ("D2","E4 Matching & LLM","Blended fuzzy scorer","0.45 token-set + 0.30 token-sort + 0.25 soft-Jaccard.","Must",5,"S0","Done","D1","Single-metric gave a false 100%"),
 ("D3","E4 Matching & LLM","Alias learning","An override is stored as an alias so the same wording resolves next time without an LLM call.","Must",3,"S0","Done","D1",""),
 ("D4","E4 Matching & LLM","LLM provider abstraction","One config line switches anthropic / gemini / openai / ollama / none.","Must",5,"S0","Done","","Scaffold complete"),
 ("D9","E4 Matching & LLM","Invert to LLM-first matching","The LLM decides group then specs on every judgement call; the rules produce the shortlist and constraints and hold the veto. Every decision stamped matched_by.","Must",8,"R3","Not started","D4","Your instruction - reverses the earlier design"),
 ("D10","E4 Matching & LLM","Fallback on rate limit or quota","429, timeout, bad key or no network falls straight through to the deterministic result, marked matched_by:rules. The tool never stops working.","Must",5,"R3","Not started","D9","The always-there fallback"),
 ("D11","E4 Matching & LLM","Cache and batch LLM calls","Results cached on normalised text; a 20-line invoice resolves in one call, not twenty. Exact phase-1 hits never call out.","Must",5,"R4","Not started","D9","Keeps LLM-first affordable"),
 ("D7","E4 Matching & LLM","LLM guardrails","The model may only pick from the shortlist or answer none. It can never mint a code.","Must",3,"S0","Done","D4",""),
 ("D8","E4 Matching & LLM","Graceful degradation","Provider down or key invalid falls back to fuzzy plus a human question, never a crash.","Must",2,"R5","Not started","D4",""),
 ("F1","E5 Invoice & OCR","Multi-format extraction","Text, PDF, scanned PDF, image, Excel and CSV all produce a line list.","Must",8,"S0","Done","",""),
 ("F2","E5 Invoice & OCR","Extraction cascade","Tables, then text layer, then OCR per page. Only pages without text go to OCR.","Must",5,"S0","Done","F1",""),
 ("F6","E5 Invoice & OCR","Field extraction","Description, quantity, UoM, rate, HSN and line-level vendor. Header supplier ignored.","Must",5,"S0","Done","F1","Your rule: line vendor only"),
 ("F7","E5 Invoice & OCR","Independent batch handling","A 20-line invoice yields 20 proposals; one failure never blocks the rest.","Must",3,"S0","Done","F1",""),
 ("G1","E6 Master & Restructure","Item master screen","Search, filter and edit, attributed and logged.","Must",5,"S0","Done","",""),
 ("G2","E6 Master & Restructure","Field edits never touch the code","Enforced in the API, not only the interface.","Must",2,"S0","Done","G1",""),
 ("G4","E6 Master & Restructure","Code mapping ledger","Every old-to-new pair recorded and exported.","Must",3,"S0","Done","G3",""),
 ("H1","E7 ERPNext","REST client","Login, retry and error surfacing against UAT and PROD.","Must",3,"S0","Done","",""),
 ("H2","E7 ERPNext","Pull live item list","One click refreshes the reserved-code set from live ERPNext.","Must",3,"S0","Done","H1",""),
 ("H3","E7 ERPNext","Dry-run payload preview","Shows the exact JSON that would be posted, writes nothing.","Must",3,"S0","Done","H1",""),
 # ---------------------------------------------------------------- NEEDS YOUR INPUT
 ("D5","E4 Matching & LLM","Choose the LLM provider and wire the key","A provider is chosen, the key stored outside the repo, per-call cost logged.","Must",3,"NEEDS INPUT","Blocked","D4","Needs: an API key"),
 ("D6","E4 Matching & LLM","Calibrate the 60% threshold","200 real invoice lines run; false accepts and false asks measured; the chosen number recorded.","Must",8,"NEEDS INPUT","Blocked","D5","Needs: 200 real lines + a key"),
 ("F3","E5 Invoice & OCR","Install Tesseract on the host","OCR verified working on the machine that will host the app.","Must",2,"NEEDS INPUT","Blocked","J7","Needs: admin rights on the host"),
 ("F4","E5 Invoice & OCR","OCR accuracy on real invoices","Ten real scanned invoices processed; per-field accuracy recorded; failures listed.","Must",5,"NEEDS INPUT","Blocked","F3","Needs: 10 real scanned invoices"),
 ("F5","E5 Invoice & OCR","Line filter tuning","GST rows, totals, bank details and signatures excluded on the real sample.","Must",3,"NEEDS INPUT","Blocked","F4","Needs: the same invoices"),
 ("H9","E7 ERPNext","Service account and role","A dedicated itemcode.studio user with an Item Code Studio role, API key not password, delete refused everywhere, transactions and permission doctypes refused.","Must",5,"NEEDS INPUT","Blocked","","Needs: someone with ERP admin rights"),
 ("H10","E7 ERPNext","Prove the guardrail","The service account attempts one write to Stock Entry and one to Custom Field. Both must fail with 403.","Must",3,"NEEDS INPUT","Blocked","H9","A guardrail nobody has tried is not a guardrail"),
 ("H11","E7 ERPNext","Sub-head to Item Group mapping","Each of our 46 sub-heads maps to one of ERPNext 23 item groups. Mapping is data, not code, and is editable.","Must",5,"R5","Not started","","ERP has 23 groups, we have 889"),
 ("H12","E7 ERPNext","Validate UoM and HSN against ERP","Stock UoM checked against the 240 UOM records and HSN against the 18,689 GST HSN Code records before any write.","Must",3,"R5","Not started","H11","Stops bad values at source"),
 ("H13","E7 ERPNext","Populate the ERP specification fields","item_specification_1-4 and item_vendor written on create so the decode lives in ERPNext too.","Should",5,"NEEDS INPUT","Blocked","H4","Needs: decision on Wahni parallel framework"),
 ("H14","E7 ERPNext","Twice-daily sync with drift detection","Runs at sync.times and on demand. Refreshes erp_item, reports items only in ERP, only local, and disagreeing fields. Conflicts are surfaced, never auto-resolved.","Must",8,"R5","Not started","H12","New requirement"),
 ("H4","E7 ERPNext","Live create on UAT","Submitting creates the Item in UAT with our code as item_code.","Must",5,"NEEDS INPUT","Blocked","H3","Needs: your go-ahead to write"),
 ("H5","E7 ERPNext","Per-creator ERP credentials","Each creator authenticates to ERPNext as themselves.","Must",5,"NEEDS INPUT","Blocked","A5","Needs: ERP logins per person"),
 ("H6","E7 ERPNext","Rename cascade","When an unfrozen item re-codes, the ERPNext rename is issued and confirmed.","Must",5,"NEEDS INPUT","Blocked","H4","Needs: UAT writes first"),
 ("H7","E7 ERPNext","Drift reconciliation report","Lists app-only, ERP-only and disagreeing fields.","Should",5,"NEEDS INPUT","Blocked","H2","Needs: UAT writes first"),
 ("H8","E7 ERPNext","PROD go-live","Writes enabled on PROD after a clean UAT week. Rollback path written down.","Must",3,"NEEDS INPUT","Blocked","H4;H5","Needs: your sign-off"),
 ("I1","E8 Data Quality","Resolve 10 unmatched items","The 10 seeded items with no matching group are assigned or corrected.","Must",2,"NEEDS INPUT","Blocked","","Needs: 10 judgement calls from you"),
 ("I2","E8 Data Quality","Triage 175 junk ERP codes","LCO-1, IBC TANKS, Iron and bare numbers each disabled or re-coded.","Should",5,"NEEDS INPUT","Blocked","H4","Needs: ERP write access"),
 ("I3","E8 Data Quality","Policy for 614 non-grammar codes","Leave frozen, or plan a one-time re-code. Written either way.","Must",3,"NEEDS INPUT","Blocked","","Needs: a business decision"),
 ("I4","E8 Data Quality","Group de-duplication campaign","889 groups reviewed, duplicates merged, an agreed target reached.","Must",13,"NEEDS INPUT","Blocked","G3","Needs: your judgement on 889 groups"),
 ("I5","E8 Data Quality","Spec value de-duplication","Within-group duplicate values merged, aliases preserved.","Should",5,"NEEDS INPUT","Blocked","I4","Needs: I4 first"),
 ("I6","E8 Data Quality","HSN backfill","Blank and 4-digit HSN filled where a correct 8-digit code exists.","Should",5,"NEEDS INPUT","Blocked","","Needs: CA verification"),
 ("J1","E9 Deploy & Handover","Nominate the host machine","An always-on machine, sleep disabled, fixed hostname.","Must",2,"NEEDS INPUT","Blocked","","Needs: which machine"),
 ("J4","E9 Deploy & Handover","Operator SOP and training","One-page sheet plus a session with the creators.","Must",3,"NEEDS INPUT","Blocked","J7","Text drafted in PDR §8"),
 ("J5","E9 Deploy & Handover","Support runbook","Restart, restore, add a user, rotate a key, who to call.","Must",3,"NEEDS INPUT","Blocked","J2",""),
 ("J6","E9 Deploy & Handover","Two-week post-launch watch","Daily check of codes issued, questions raised and mismatches, with a written close-out.","Must",5,"NEEDS INPUT","Blocked","H8","Needs: two weeks to pass"),
 ("G6","E6 Master & Restructure","Bulk operations","Select many items and move or edit them in one previewed action.","Could",5,"BACKLOG","Not started","G3","Deferred, not needed for go-live"),
 ("C9","E3 Code Engine","Vacancy claim override","An admin can release or reassign a parked number, with a reason recorded.","Could",3,"BACKLOG","Not started","C4","Deferred"),
]

# ---------------------------------------------------------------- Delivery Plan
ws = wb.active; ws.title = "Delivery Plan"
ws["A1"] = "Item Code Studio — Delivery Plan"; ws["A1"].font = TITLE
ws["A2"] = ("Prepared for Anuraag · 6 August 2026 · v2.0 · R1-R5 are today's build runs.  "
            "S0 = already built and verified.  NEEDS INPUT = cannot be done by me; the Notes column says what is required.")
ws["A2"].font = S
cols = [("ID",7),("Agent",14),("Epic",24),("Story",36),("Acceptance criteria",64),("Priority",9),
        ("Points",8),("Run",13),("Status",12),("Depends on",12),("Notes",34)]
r0 = 4
for i,(h,w) in enumerate(cols,1):
    c = ws.cell(r0,i,h); c.font = H; c.fill = PatternFill("solid",fgColor=NAVY)
    c.alignment = CTR; c.border = BORD
    ws.column_dimensions[get_column_letter(i)].width = w
for j,row in enumerate(ST):
    r = r0+1+j
    vals = (row[0], AGENT.get(row[0],"?")) + row[1:]
    for i,v in enumerate(vals,1):
        c = ws.cell(r,i,v); c.font = N; c.border = BORD
        c.alignment = CTR if i in (1,2,6,7,8,9,10) else WRAP
    if row[7] == "Done":      fill = GREEN
    elif row[7] == "Blocked": fill = AMBER
    elif row[6] == "BACKLOG": fill = LGREY
    else:                     fill = ICE if j % 2 else None
    if fill:
        for i in range(1,12): ws.cell(r,i).fill = PatternFill("solid",fgColor=fill)
    ws.row_dimensions[r].height = 30
last = r0+len(ST)
tr = last+1
ws.cell(tr,4,"TOTAL").font = B
ws.cell(tr,7,f"=SUM(G{r0+1}:G{last})").font = B
for i in range(1,12):
    ws.cell(tr,i).fill = PatternFill("solid",fgColor=LGREY); ws.cell(tr,i).border = BORD
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A{r0}:K{last}"
dv = DataValidation(type="list", formula1='"Not started,In progress,Blocked,Done"', allow_blank=True)
ws.add_data_validation(dv); dv.add(f"I{r0+1}:I{last}")

# ---------------------------------------------------------------- Runs
sp = wb.create_sheet("Runs")
sp["A1"] = "Build runs — one sitting, today"; sp["A1"].font = TITLE
sp["A2"] = "Points roll up from the Delivery Plan. Each run ends with something you can open and use."
sp["A2"].font = S
sh = ["Run","What it delivers","Points","Stories","Done","Complete %"]
w = [13,74,10,10,9,12]
for i,(h,wd) in enumerate(zip(sh,w),1):
    c = sp.cell(4,i,h); c.font = H; c.fill = PatternFill("solid",fgColor=STEEL)
    c.alignment = CTR; c.border = BORD
    sp.column_dimensions[get_column_letter(i)].width = wd
RUNS = [
 ("S0","Already built: the grammar, all three phases, matching, invoice extraction, the master, restructuring, export, ERP read"),
 ("R1","MiniMines theme and logo, the split landing page — decoder one side, directory the other, log in on top. Stats block removed"),
 ("R2","Login, sessions, attribution from the session, user management, the creator shell behind the gate"),
 ("R3","The create screen: paste or upload an invoice, three phases shown per line, cascading head-sub-head-group dropdowns, live code recomputation, edit before submit"),
 ("R4","Item master as a fully editable directory, revision history on every save, one-click revert, activity feed with field-level diffs"),
 ("R5","API frozen at /api/v1 with a real error contract, smoke suite, deployed to the internal server, starts on boot, daily backup"),
 ("NEEDS INPUT","Blocked on something only you or the organisation can supply — see the next tab"),
 ("BACKLOG","Agreed but not needed for go-live"),
]
for j,(rn,d) in enumerate(RUNS):
    r = 5+j
    sp.cell(r,1,rn).font = B; sp.cell(r,2,d).font = N
    sp.cell(r,3,f"=SUMIFS('Delivery Plan'!$G:$G,'Delivery Plan'!$H:$H,$A{r})")
    sp.cell(r,4,f"=COUNTIFS('Delivery Plan'!$H:$H,$A{r})")
    sp.cell(r,5,f"=SUMIFS('Delivery Plan'!$G:$G,'Delivery Plan'!$H:$H,$A{r},'Delivery Plan'!$I:$I,\"Done\")")
    sp.cell(r,6,f"=IFERROR(E{r}/C{r},0)")
    sp.cell(r,6).number_format = "0%"
    for i in range(1,7):
        sp.cell(r,i).border = BORD; sp.cell(r,i).font = N if i>1 else B
        sp.cell(r,i).alignment = WRAP if i==2 else CTR
    sp.row_dimensions[r].height = 34
    if rn == "S0":           f_ = DONEG
    elif rn == "NEEDS INPUT":f_ = AMBER
    elif rn == "BACKLOG":    f_ = LGREY
    else:                    f_ = None
    if f_:
        for i in range(1,7): sp.cell(r,i).fill = PatternFill("solid",fgColor=f_)
tr2 = 5+len(RUNS)
sp.cell(tr2,2,"TOTAL").font = B
sp.cell(tr2,3,f"=SUM(C5:C{tr2-1})").font = B
sp.cell(tr2,4,f"=SUM(D5:D{tr2-1})").font = B
sp.cell(tr2,5,f"=SUM(E5:E{tr2-1})").font = B
sp.cell(tr2,6,f"=IFERROR(E{tr2}/C{tr2},0)").font = B
sp.cell(tr2,6).number_format = "0%"
for i in range(1,7):
    sp.cell(tr2,i).fill = PatternFill("solid",fgColor=LGREY); sp.cell(tr2,i).border = BORD
    sp.cell(tr2,i).alignment = CTR
sp.cell(tr2+2,1,"Buildable today (R1-R5):").font = B
sp.cell(tr2+2,3,f"=SUMIFS('Delivery Plan'!$G:$G,'Delivery Plan'!$G:$G,\"R1\")"
                f"+SUMIFS('Delivery Plan'!$G:$G,'Delivery Plan'!$G:$G,\"R2\")"
                f"+SUMIFS('Delivery Plan'!$G:$G,'Delivery Plan'!$G:$G,\"R3\")"
                f"+SUMIFS('Delivery Plan'!$G:$G,'Delivery Plan'!$G:$G,\"R4\")"
                f"+SUMIFS('Delivery Plan'!$G:$G,'Delivery Plan'!$G:$G,\"R5\")").font = B
sp.cell(tr2+3,1,"Blocked on your input:").font = B
sp.cell(tr2+3,3,f"=SUMIFS('Delivery Plan'!$G:$G,'Delivery Plan'!$G:$G,\"NEEDS INPUT\")").font = B
sp.cell(tr2+5,1,"Every run leaves the app working. If we stop after R3 you still have a usable tool, just without version control.").font = S

# ---------------------------------------------------------------- Needs input
ni = wb.create_sheet("Needs Your Input")
ni["A1"] = "What I cannot do for you"; ni["A1"].font = TITLE
ni["A2"] = ("These are real work items, not excuses. Each needs a thing only you or the organisation can supply. "
            "Give me the input and the work moves in the same sitting.")
ni["A2"].font = S
nh = ["What I need from you","Unblocks","Points","Why it cannot be faked"]
nw = [42,40,10,64]
for i,(h,wd) in enumerate(zip(nh,nw),1):
    c = ni.cell(4,i,h); c.font = H; c.fill = PatternFill("solid",fgColor=STEEL)
    c.alignment = CTR; c.border = BORD
    ni.column_dimensions[get_column_letter(i)].width = wd
NEED = [
 ("An LLM API key (Gemini Flash is free at this volume)","D5, D6 — the sub-60% matching layer",11,
  "Without a provider the third matching layer never runs. Fuzzy alone still works, but the hard lines stay questions for the operator."),
 ("Ten real scanned or photographed invoices","F3, F4, F5 — OCR proven end to end",10,
  "OCR accuracy depends entirely on your paper: fonts, stamps, skew, scan quality. Any number I quote from synthetic files would be fiction."),
 ("Written go-ahead to write into ERPNext UAT","H4, H6, H7 — items actually created",15,
  "Writing to a live system is your call, not mine. Dry-run shows the exact payload first."),
 ("An ERPNext login per creator","H5, H8 — PROD go-live",8,
  "A shared account leaves an audit trail naming nobody. This blocks PROD by design, not by preference."),
 ("Which desktop or box hosts it, and can it stay on","J1, F3, J7 — deployment",9,
  "Everything else is ready. I need a machine name and admin rights on it."),
 ("Your judgement on 889 groups: how far to cut","I1, I3, I4, I5 — the clean-up",23,
  "Which groups are genuinely duplicates is a business judgement about your materials. Guessing would corrupt the master."),
 ("CA verification of specialised HSN codes","I6 — HSN backfill",5,
  "Carried over from the item master work. Wrong HSN is a GST exposure, so it needs a human who is accountable."),
 ("Two weeks after go-live, and the creators' time","J4, J5, J6 — training and the watch",11,
  "Training needs people in a room. The watch needs time to pass."),
]
for j,row in enumerate(NEED):
    r = 5+j
    for i,v in enumerate(row,1):
        c = ni.cell(r,i,v); c.font = N; c.border = BORD
        c.alignment = CTR if i==3 else WRAP
    ni.row_dimensions[r].height = 42
tr4 = 5+len(NEED)
ni.cell(tr4,2,"TOTAL BLOCKED").font = B
ni.cell(tr4,3,f"=SUM(C5:C{tr4-1})").font = B
for i in range(1,5):
    ni.cell(tr4,i).fill = PatternFill("solid",fgColor=AMBER); ni.cell(tr4,i).border = BORD
    ni.cell(tr4,i).alignment = CTR

# ---------------------------------------------------------------- Epics
ep = wb.create_sheet("Epics")
ep["A1"] = "Epics"; ep["A1"].font = TITLE
ep["A2"] = "Calculated from the Delivery Plan."; ep["A2"].font = S
eh = ["Epic","What it covers","Stories","Points","Done pts","Complete %"]
ew = [24,64,10,10,11,12]
for i,(h,wd) in enumerate(zip(eh,ew),1):
    c = ep.cell(4,i,h); c.font = H; c.fill = PatternFill("solid",fgColor=STEEL)
    c.alignment = CTR; c.border = BORD
    ep.column_dimensions[get_column_letter(i)].width = wd
EP = [
 ("E0 Brand & Shell","MiniMines palette, logo and typeface; the stats block removed"),
 ("E1 Access & Identity","The split landing page anyone can use, and a login that leads to creation"),
 ("E2 API v1","A stable, documented, versioned contract for the larger system"),
 ("E3 Code Engine","The grammar, three phases, numbering, vacancies, freeze-on-first-use"),
 ("E4 Matching & LLM","Normaliser, fuzzy blend, alias learning, LLM below 60% confidence"),
 ("E5 Invoice & OCR","Reading text, PDF, scanned and photographed invoices into resolvable lines"),
 ("E6 Master & Restructure","Editable master, cascading dropdowns, revision history, revert, restructuring"),
 ("E7 ERPNext","Reading live truth, and writing items back on submit"),
 ("E8 Data Quality","Duplicate groups, junk codes, missing HSN"),
 ("E9 Deploy & Handover","Server, backup, start-on-boot, training, runbook, post-launch watch"),
]
for j,(e,d) in enumerate(EP):
    r = 5+j
    ep.cell(r,1,e).font = B; ep.cell(r,2,d).font = N
    ep.cell(r,3,f"=COUNTIFS('Delivery Plan'!$C:$C,$A{r})")
    ep.cell(r,4,f"=SUMIFS('Delivery Plan'!$G:$G,'Delivery Plan'!$C:$C,$A{r})")
    ep.cell(r,5,f"=SUMIFS('Delivery Plan'!$G:$G,'Delivery Plan'!$C:$C,$A{r},'Delivery Plan'!$I:$I,\"Done\")")
    ep.cell(r,6,f"=IFERROR(E{r}/D{r},0)")
    ep.cell(r,6).number_format = "0%"
    for i in range(1,7):
        ep.cell(r,i).border = BORD
        ep.cell(r,i).alignment = WRAP if i in (1,2) else CTR
        if i>2: ep.cell(r,i).font = N
    ep.row_dimensions[r].height = 28
tr3 = 5+len(EP)
ep.cell(tr3,2,"TOTAL").font = B
for i,f_ in [(3,f"=SUM(C5:C{tr3-1})"),(4,f"=SUM(D5:D{tr3-1})"),(5,f"=SUM(E5:E{tr3-1})")]:
    ep.cell(tr3,i,f_).font = B
ep.cell(tr3,6,f"=IFERROR(E{tr3}/D{tr3},0)").font = B
ep.cell(tr3,6).number_format = "0%"
for i in range(1,7):
    ep.cell(tr3,i).fill = PatternFill("solid",fgColor=LGREY); ep.cell(tr3,i).border = BORD
    ep.cell(tr3,i).alignment = CTR

# ---------------------------------------------------------------- Risks
rk = wb.create_sheet("Risks")
rk["A1"] = "Risks"; rk["A1"].font = TITLE
rh = ["#","Risk","Impact","Likelihood","Response","Status"]
rw = [5,54,11,12,60,13]
for i,(h,wd) in enumerate(zip(rh,rw),1):
    c = rk.cell(3,i,h); c.font = H; c.fill = PatternFill("solid",fgColor=STEEL)
    c.alignment = CTR; c.border = BORD
    rk.column_dimensions[get_column_letter(i)].width = wd
RISK = [
 (1,"Building all five runs in one sitting means less testing per feature than the six-sprint plan allowed.","High","Certain",
  "Each run ends working and is exercised against the real seeded data before the next starts. Where a run is thin, it is named in the handover rather than hidden.","Accepted"),
 (2,"Version control plus freeze-on-first-use interact awkwardly: a revert cannot restore a code that ERPNext has frozen.","High","Likely",
  "K10 makes revert restore fields only, and say plainly what it could not restore. Never a silent partial revert.","Open"),
 (3,"OCR accuracy is completely unmeasured on your paper.","High","Possible",
  "Blocked on ten real invoices. Until then OCR is present but unproven, and the handover says so.","Blocked"),
 (4,"The host is a single point of failure. If it sleeps, nobody can issue a code.","High","Likely",
  "Daily off-machine backup, start-on-boot. Move to an always-on box if it bites.","Open"),
 (5,"Passwords cross the LAN in the clear over plain HTTP.","Medium","Certain",
  "Accepted for an internal tool. Use passwords not reused elsewhere. TLS if it ever leaves the LAN.","Accepted"),
 (6,"Group de-duplication keeps slipping. Every month more codes freeze and the clean-up gets permanently dearer.","High","Likely",
  "Blocked on your judgement. This is the item I would push hardest to unblock first.","Blocked"),
 (7,"Writing to PROD under a shared account leaves an audit trail naming nobody.","High","Certain",
  "Per-creator ERP credentials block PROD go-live by design.","Blocked"),
 (8,"Anyone who can log in can restructure the master, with no approval gate, by instruction.","Medium","Possible",
  "Preview, attribution, full revision history and one-click revert. Revisit only if it actually causes harm.","Accepted"),
 (9,"A structural change re-codes an item ERPNext has already transacted.","High","Unlikely",
  "Freeze-on-first-use makes it impossible by design; every change is previewed first.","Mitigated"),
 (10,"The larger system builds against an API that then shifts underneath it.","Medium","Possible",
  "R5 freezes the contract at /api/v1 before anyone integrates.","Open"),
]
for j,row in enumerate(RISK):
    r = 4+j
    for i,v in enumerate(row,1):
        c = rk.cell(r,i,v); c.font = N; c.border = BORD
        c.alignment = WRAP if i in (2,5) else CTR
    rk.cell(r,3).fill = PatternFill("solid",fgColor=RED if row[2]=="High" else AMBER)
    rk.row_dimensions[r].height = 36

# ---------------------------------------------------------------- Decisions
dc = wb.create_sheet("Decisions")
dc["A1"] = "Decision log"; dc["A1"].font = TITLE
dc["A2"] = "Amber rows need your word before or during the build."; dc["A2"].font = S
dh = ["#","Decision","Rationale","Status"]
dw = [5,62,70,20]
for i,(h,wd) in enumerate(zip(dh,dw),1):
    c = dc.cell(4,i,h); c.font = H; c.fill = PatternFill("solid",fgColor=STEEL)
    c.alignment = CTR; c.border = BORD
    dc.column_dimensions[get_column_letter(i)].width = wd
DEC = [
 (1,"Variable-length codes, never padded","Interior gaps carry meaning as 00; trailing gaps carry none. Lengths 7/9/11/13/15/17.","Locked"),
 (2,"Vendor sits last, only in groups that declare it","Four battery groups declare a Maker slot. Elsewhere one item keeps one code regardless of supplier.","Locked"),
 (3,"Vendor from the invoice line only, never the header","A supplier named at the top of a page says nothing about an individual line.","Locked"),
 (4,"No -R / -W / -F state suffixes","Removed entirely.","Locked"),
 (5,"Spec slot meaning is per group","Air Freshener uses Type and Size; Battery Pack uses Form Factor, Chemistry, Size, Capacity, Maker.","Locked"),
 (6,"Group numbers restart inside each sub-head","999 per sub-head is ample at 889 groups.","Locked"),
 (7,"A vacated group number is parked, not reused","Released only to a >=88% name match. The number carries the meaning of what held it.","Locked"),
 (8,"Prefix collisions walk to the next free combination","Seeded prefixes read from codes already issued, so CEEU and its like survive.","Locked"),
 (9,"No approval workflow","Preview, attribution, revision history and revert replace permission.","Locked"),
 (10,"Two tiers, not roles: public read, login to create","You have two kinds of people and one permission. A role matrix would never earn its keep.","Locked"),
 (11,"Landing page is split: decoder one side, directory the other","Most people only ever need to look something up. That is the front door.","Locked"),
 (12,"Item master is the editable twin of the public directory","Same rows, same columns, editable behind the login.","Locked"),
 (13,"Every save is a version, and any version can be restored","Replaces approval with reversibility.","Locked"),
 (14,"Simple username and password, no OTP, no SSO","Read from 'simple authentication'. If you meant a one-time code by email or SMS, say so — it needs a delivery channel and is a different build.","For confirmation"),
 (15,"LLM-first matching, rules as guardrail and fallback","The LLM decides group then specs; the rules build the shortlist, hold the veto, and take over entirely on rate limit or outage.","Locked"),
 (19,"Vacated numbers are claimed by matching, not by queue position","A reserved number waits for something that genuinely matches what left, rather than going to the next arrival.","For confirmation"),
 (20,"Populate Wahni's ERPNext specification fields","ERP already has item_specification_1-4, item_vendor and 62 spec records, barely used. Populate them rather than leave two parallel systems.","For confirmation"),
 (21,"Map our taxonomy onto ERP Item Groups at sub-head level","ERP has 23 item groups, we have 889. Creating 889 would wreck their reporting tree.","For confirmation"),
 (16,"Freeze on first use: codes live in ERPNext are never rewritten","Submitted Frappe documents are immutable and renames cascade. Kept and flagged stale instead.","For confirmation"),
 (17,"A revert restores fields, never a frozen code","Where version control meets ERPNext reality. It says plainly what it could not restore.","For confirmation"),
 (18,"One SQLite file on one internal server","Makes duplicate codes structurally impossible. Accepts the host as a single point of failure.","Locked"),
]
for j,row in enumerate(DEC):
    r = 5+j
    for i,v in enumerate(row,1):
        c = dc.cell(r,i,v); c.font = N; c.border = BORD
        c.alignment = WRAP if i in (2,3) else CTR
    if "confirmation" in str(row[3]):
        for i in range(1,5): dc.cell(r,i).fill = PatternFill("solid",fgColor=AMBER)
    dc.row_dimensions[r].height = 30

# ---------------------------------------------------------------- DoD
dd = wb.create_sheet("Definition of Done")
dd["A1"] = "Definition of Done"; dd["A1"].font = TITLE
dd["A2"] = "Compressed to one sitting, so this is the floor, not the ceiling."; dd["A2"].font = S
dd.column_dimensions["A"].width = 5
dd.column_dimensions["B"].width = 96
DOD = [
 "The acceptance criteria on the Delivery Plan row are met as written, not as reinterpreted.",
 "Exercised against the real seeded data — 889 groups, 1,947 items, 2,677 live ERP codes — never a toy fixture.",
 "Any rule that protects the master is enforced in the API, not only in the interface.",
 "One failure path tried on purpose, producing a clear message rather than a crash or a silent wrong answer.",
 "Every write is attributed to the authenticated user and appears in the activity log.",
 "The app still starts and the previous runs still work.",
 "Anything thin or untested is written into the handover, not carried in someone's head.",
]
for j,t in enumerate(DOD):
    r = 4+j
    dd.cell(r,1,j+1).font = B; dd.cell(r,1).alignment = CTR
    c = dd.cell(r,2,t); c.font = N; c.alignment = WRAP
    dd.row_dimensions[r].height = 26

for s in wb.worksheets:
    s.sheet_view.showGridLines = False

wb.save(r"C:\Users\Anura\ItemCodeStudio\Sprint_Plan_v3.xlsx")
print("written")
