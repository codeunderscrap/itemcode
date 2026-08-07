"""Excel export - the item master in the same shape the business already uses."""
import json
import os

from .db import now

MASTER_COLS = ["Item Code", "Item Name", "Head", "Sub Head", "Item Group",
               "Spec 1", "Spec 2", "Spec 3", "Spec 4", "Vendor",
               "Description", "UoM", "Alternate UoM", "HSN/SAC", "Item Tax Template",
               "Maintain Stock", "Allow Sales", "Has Batch No",
               "Status", "In ERPNext", "Decodable", "Created By", "Created On"]

ICM_COLS = ["SL", "Item Head", "Item Sub Head", "Item Group", "Code",
            "Spec 1", "Code.1", "Spec 2", "Code.2", "Spec 3", "Code.3",
            "Spec 4", "Code.4", "Vendor", "Code.5", "Final Code", "UoM"]


def _sv(con, sid):
    if not sid:
        return None
    r = con.execute("SELECT value FROM specval WHERE id=?", (sid,)).fetchone()
    return r["value"] if r else None


def export(con, out_dir, filename=None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    grp_fill = PatternFill("solid", fgColor="DDEBF7")
    warn_fill = PatternFill("solid", fgColor="FCE4D6")

    def head_row(ws, cols):
        ws.append(cols)
        for i in range(1, len(cols) + 1):
            c = ws.cell(row=1, column=i)
            c.font, c.fill = hdr_font, hdr_fill
            c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"

    # ---------------------------------------------------------- Item_Master
    ws = wb.active
    ws.title = "Item_Master"
    head_row(ws, MASTER_COLS)
    rows = con.execute("""
        SELECT i.*, g.name AS gname, s.name AS sname, h.name AS hname
        FROM item i LEFT JOIN grp g ON g.id=i.grp_id
        LEFT JOIN subhead s ON s.id=g.subhead_id LEFT JOIN head h ON h.id=s.head_id
        ORDER BY i.code""").fetchall()
    for r in rows:
        ws.append([r["code"], r["name"], r["hname"], r["sname"], r["gname"],
                   _sv(con, r["s1"]), _sv(con, r["s2"]), _sv(con, r["s3"]), _sv(con, r["s4"]),
                   _sv(con, r["vend"]), r["description"], r["uom"], r["alt_uom"], r["hsn"],
                   r["tax"], r["maintain_stock"], r["allow_sales"], r["has_batch"],
                   r["status"], "Yes" if r["status"] == "in_erp" else "No",
                   "Yes" if r["decodable"] else "No", r["created_by"], r["created_at"]])
        if not r["decodable"]:
            for i in range(1, len(MASTER_COLS) + 1):
                ws.cell(row=ws.max_row, column=i).fill = warn_fill

    # ------------------------------------------------------ Item_Code_Master
    ws2 = wb.create_sheet("Item_Code_Master")
    head_row(ws2, ICM_COLS)
    sl = 0
    for g in con.execute("""
            SELECT g.*, s.name AS sname, s.code2 AS scode, h.name AS hname, h.code2 AS hcode
            FROM grp g JOIN subhead s ON s.id=g.subhead_id JOIN head h ON h.id=s.head_id
            WHERE g.status='active' ORDER BY h.code2, s.code2, g.code3"""):
        sl += 1
        labels = json.loads(g["labels"] or "{}")
        ws2.append([sl, g["hname"], g["sname"], g["name"], g["code3"],
                    labels.get("1"), "", labels.get("2"), "", labels.get("3"), "",
                    labels.get("4"), "", labels.get("vendor"), "",
                    f"{g['hcode']}{g['scode']}{g['code3']}", g["uom"]])
        for i in range(1, len(ICM_COLS) + 1):
            ws2.cell(row=ws2.max_row, column=i).fill = grp_fill
            ws2.cell(row=ws2.max_row, column=i).font = Font(bold=True)
        vals = {s: [dict(r) for r in con.execute(
            "SELECT value,code2 FROM specval WHERE grp_id=? AND slot=? ORDER BY code2",
            (g["id"], s))] for s in (1, 2, 3, 4, 5)}
        for i in range(max((len(v) for v in vals.values()), default=0)):
            row = ["", "", "", "", ""]
            for s in (1, 2, 3, 4, 5):
                v = vals[s][i] if i < len(vals[s]) else None
                row += [v["value"] if v else "", v["code2"] if v else ""]
            row += ["", ""]
            ws2.append(row)

    # ------------------------------------------------------- Code_Mapping
    ws3 = wb.create_sheet("Code_Mapping")
    head_row(ws3, ["Old Code", "New Code", "Reason", "Changed By", "When", "Pushed to ERP"])
    for r in con.execute("SELECT * FROM code_mapping ORDER BY ts DESC"):
        ws3.append([r["old_code"], r["new_code"], r["reason"], r["user"], r["ts"],
                    "Yes" if r["pushed_to_erp"] else "No"])

    # ------------------------------------------------------------- Audit
    ws4 = wb.create_sheet("Audit_Trail")
    head_row(ws4, ["When", "User", "Action", "Target", "Detail"])
    for r in con.execute("SELECT * FROM audit ORDER BY id DESC LIMIT 5000"):
        ws4.append([r["ts"], r["user"], r["action"], r["target"], (r["detail"] or "")[:500]])

    for sheet in wb.worksheets:
        for col in range(1, sheet.max_column + 1):
            width = max((len(str(sheet.cell(row=rr, column=col).value or ""))
                         for rr in range(1, min(sheet.max_row, 300) + 1)), default=10)
            sheet.column_dimensions[get_column_letter(col)].width = min(max(width + 2, 10), 46)

    os.makedirs(out_dir, exist_ok=True)
    fn = filename or f"Item_Master_{now().replace(':', '').replace('-', '')}.xlsx"
    path = os.path.join(out_dir, fn)
    wb.save(path)
    return path
